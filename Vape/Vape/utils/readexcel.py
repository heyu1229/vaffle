import ast
import json

from openpyxl import load_workbook

from utils.getrandom import rand
from vape.login import signin
from vape.request import vapePost, KEY

'''
@create : derrick
@Date :2019/9/19
@desc :
'''


class excel:
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.sheetnames
        self.sheet = sheets[2]
        self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def getCellValue(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def setCellValue(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

    def checkRsult(self, result, i, code):
        try:
            if str(json.loads(result)['code']) == code:
                self.setCellValue(i, 9, json.loads(result)['code'])
                self.setCellValue(i, 10, 'success')
                self.setCellValue(i, 11, json.loads(result)['runtime'])
            else:
                self.setCellValue(i, 9, result)
                self.setCellValue(i, 10, 'Faile')
        except:
            self.setCellValue(i, 9, result)
            self.setCellValue(i, 10, 'Faile')

    def readExcel(self):
        print('---------------start------------------')
        params = 'params'
        for i in range(2, self.ws.max_row + 1):
            isRun = self.getCellValue(i, 8)
            param = self.getCellValue(i, 2)
            if param is None:
                param = params
            else:
                param = ast.literal_eval(self.getCellValue(i, 2))
            if isRun == 'N':
                continue
            else:
                name = self.getCellValue(i, 3)
                pwd = str(self.getCellValue(i, 4))
                address = self.getCellValue(i, 5)
                code = self.getCellValue(i, 6)
                serial = rand()
                if param != params and name is not None:
                    login = signin(name, pwd, serial)
                    mix_secret = json.loads(login)['data']['mix_secret']
                    uuid = json.loads(login)['data']['uuid']
                    result = vapePost(address, mix_secret, serial, params=param, uuid=uuid)
                    self.checkRsult(result, i, code)
                elif param != params and name is None:
                    result = vapePost(address, KEY, serial, params=param)
                    self.checkRsult(result, i, code)
                elif param == params and name is not None:
                    login = signin(name, pwd, serial)
                    mix_secret = json.loads(login)['data']['mix_secret']
                    uuid = json.loads(login)['data']['uuid']
                    result = vapePost(address, mix_secret, serial, uuid=uuid)
                    self.checkRsult(result, i, code)
                elif param == params and name is None:
                    result = vapePost(address, KEY, serial)
                    self.checkRsult(result, i, code)
        print('---------------end------------------')


e = excel('vaffle.xlsx')
e.readExcel()
# t = readExcel('vaffle.xlsx', 'vaffle')
